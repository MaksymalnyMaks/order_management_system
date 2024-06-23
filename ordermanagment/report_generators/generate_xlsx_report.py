from openpyxl import Workbook
from openpyxl.styles import PatternFill


class XlsxReportGenerator:

    def __init__(self, data: list, columns_headers: list):
        self.column_headers = columns_headers
        self.data = data

    def generate_orders_report(self, status_colors: dict):
        # Create a workbook with 'orders report' worksheet
        workbook = Workbook()
        worksheet = workbook.create_sheet(title='orders report')
        orders_report_sheet = workbook.get_sheet_by_name('orders report')
        workbook.active = orders_report_sheet

        # Add columns' headers to the worksheet
        worksheet.append(self.column_headers)

        # Add orders to the worksheet
        for order in self.data:
            row = [order.order_id, order.order_name, order.description,
                   order.creation_date.strftime("%d-%m-%Y %H:%M:%S"), order.status]
            worksheet.append(row)

            # Fill row with proper color
            fill_color = status_colors.get(order.status, 'FFFFFF')
            for cell in worksheet[worksheet.max_row]:
                cell.fill = PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color)

        return workbook
